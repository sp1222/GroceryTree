from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import GroceryCategoryTreeClass
import GroceryItemsInformationClass
import json
import re
import time


# funtion where selenium gathers html from each web page
def seleniumGetsHTML(site):


    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.get(site)
    time.sleep(6)  ### Timer to allow time for the compiler to grab html

    html = BeautifulSoup(driver.page_source, 'html.parser')

    driver.close()
    driver.quit()

    return html

 

## returns the text of the starting node.
# add the first node to the tree here...
def buildTree(html, site):
    
    ### FIRST we try to find the first node of our tree, generally 'Shop' or 'Menu', or etc..
    # and we sent this first node back to main with Name of Node and href.

    startingKeyword = ['SHOP']#, 'SHOP BY DEPARTMENT', 'MENU']  # MAYBE ADD DEPARTMENT, MENU, SHOP BY DEPARTMENT, ETC..
    tagSearch = html.findAll('a')  # this will work for heb.com, will need to adjust for other websites..   
    rootNodeOfTree = GroceryCategoryTreeClass.GroceryCategoryTree()

    htmlNode = ''
    # Finds the startingKeyword as apart of startingKeywordTag,  <a> tag in this case   
    for search in tagSearch:
        for keyword in startingKeyword:
            if search.get_text().upper() == keyword:
                #here we append the first node in the category tree
                htmlNode = search  # first node's name in the tree.

    rootKey = 0
    rootNodeOfTree.setCategoryName(htmlNode.get_text())
    rootNodeOfTree.setCategoryKey(rootKey)            # root node will have a key of 0 every time.
    rootNodeOfTree.setCategoryParent(None)    # root node has no parents
    rootNodeOfTree.setCategoryParentKey(rootKey - 1)     # root node does not have a parent, thus root node's parent key is set to -1
    rootNodeOfTree.setCategory_href(htmlNode['href'])

    # root node is in the tree, we will now add categories to the tree.
    driver = webdriver.Chrome(ChromeDriverManager().install()) 
    driver.get(site)

    rootKey = addSubCategoryToCurrentCategory(rootNodeOfTree, (rootKey + 1), driver, site)
    
    driver.close()
    driver.quit()

    return rootNodeOfTree


#******************************************************************************************************************************
# Outputting tree to screen
# prints a visual representation of the categories and sub categories tree.
def printCategoryTree(currentCategory, level):

    if currentCategory.doSubCategoriesExist() == True:
        for sub in currentCategory.getSubCategories():
            print((' |    ')*level)
            print((' |    ')*level)
            print((' |    ')*(level-1) + ' |--' + sub.getCategoryName() + '  Key: ' + str(sub.getCategoryKey()))
            printCategoryTree(sub, (level+1))
        level -= 1

#*******************************************************************************************************
# Sending tree information to an excel sheet.


def createWorkbook(root, grocery):
    wb = Workbook()
    addToNewWorkbook(wb, root, 'None', -1, index = 0)
    openFile = grocery + '.xlsx'
    wb.save(openFile)

def addToNewWorkbook(wb, currentCategory, pName, pkey, index):

    wb.create_sheet(index = index, title = currentCategory.getCategoryName())
    wb.active = index
    sheet = wb.active
    catName = sheet.cell(row = 1, column = 1)
    catName.value = 'Category Name:'
    catName = sheet.cell(row = 1, column = 2)    
    catName.value = currentCategory.getCategoryName()

    catKey = sheet.cell(row = 2, column = 1)
    catKey.value = 'Category Key:'
    catKey = sheet.cell(row = 2, column = 2)    
    catKey.value = currentCategory.getCategoryKey()

    catParent = sheet.cell(row = 3, column = 1)
    catParent.value = 'Category Parent:'
    catParent = sheet.cell(row = 3, column = 2)    
    catParent.value = pName
    
    catParentKey = sheet.cell(row = 4, column = 1)
    catParentKey.value = 'Category Parent Key:'
    catParentKey = sheet.cell(row = 4, column = 2)    
    catParentKey.value = pkey
        
    catHREF = sheet.cell(row = 5, column = 1)
    catHREF.value = 'Category href:'
    catHREF = sheet.cell(row = 5, column = 2)
    catHREF.value = currentCategory.getCategory_href()

    subCatsExist = currentCategory.doSubCategoriesExist()
    catSubCatExists = sheet.cell(row = 6, column = 1)
    catSubCatExists.value = 'Sub Categories Exists:'
    catSubCatExists = sheet.cell(row = 6, column = 2)
    catSubCatExists.value = subCatsExist
    
    catSubCats = sheet.cell(row = 7, column = 1)
    catSubCats.value = 'Sub Categories:'
    catSubCatKeys = sheet.cell(row = 8, column = 1)
    catSubCatKeys.value = 'Sub Category Keys:'
    
    columnCount = 2
    for each in currentCategory.getSubCategories():
        catSubCats = sheet.cell(row = 7, column = columnCount)
        catSubCats.value = each.getCategoryName()
        catSubCatKeys = sheet.cell(row = 8, column = columnCount)
        catSubCatKeys.value = each.getCategoryKey()
        columnCount += 1
       
    label = sheet.cell(row = 10, column = 1)
    label.value = 'Name'       
    label = sheet.cell(row = 10, column = 2)
    label.value = 'Variant Price'       
    label = sheet.cell(row = 10, column = 3)
    label.value = 'Variant'       
    label = sheet.cell(row = 10, column = 4)
    label.value = 'Variant Pack Size'       
    label = sheet.cell(row = 10, column = 5)
    label.value = 'Variant Alt Price'       
    label = sheet.cell(row = 10, column = 6)
    label.value = 'Variant Alt'       
    label = sheet.cell(row = 10, column = 7)
    label.value = 'UOM Price'       
    label = sheet.cell(row = 10, column = 8)
    label.value = 'UOM'       
    label = sheet.cell(row = 10, column = 9)
    label.value = 'Brand'   
    label = sheet.cell(row = 10, column = 10)
    label.value = 'href'       
    label = sheet.cell(row = 10, column = 11)
    label.value = 'Item key'       
    label = sheet.cell(row = 10, column = 12)
    label.value = 'Category Name'       
    label = sheet.cell(row = 10, column = 13)
    label.value = 'Category Key'

    # this is where we loop each product item from currentCategory
    # and their respective product information.
    currentRow = 11
    currentColumn = 1
    for eachItem in currentCategory.getCategoryItems():
        var = sheet.cell(row = currentRow, column = currentColumn)
        var.value = eachItem.getItemName()
        currentColumn += 1
        
        var = sheet.cell(row = currentRow, column = currentColumn)
        var.value = eachItem.getItemVariantPrice()
        currentColumn += 1
        
        var = sheet.cell(row = currentRow, column = currentColumn)
        var.value = eachItem.getItemVariant()
        currentColumn += 1
        
        var = sheet.cell(row = currentRow, column = currentColumn)
        var.value = eachItem.getItemVariantPackSize()
        currentColumn += 1
        
        var = sheet.cell(row = currentRow, column = currentColumn)
        var.value = eachItem.getItemVariantAltPrice()
        currentColumn += 1
        
        var = sheet.cell(row = currentRow, column = currentColumn)
        var.value = eachItem.getItemVariantAlt()
        currentColumn += 1
        
        var = sheet.cell(row = currentRow, column = currentColumn)
        var.value = eachItem.getItemUOMPrice()
        currentColumn += 1
        
        var = sheet.cell(row = currentRow, column = currentColumn)
        var.value = eachItem.getItemUOM()
        currentColumn += 1
        
        var = sheet.cell(row = currentRow, column = currentColumn)
        var.value = eachItem.getItemBrand()
        currentColumn += 1
        
        var = sheet.cell(row = currentRow, column = currentColumn)
        var.value = eachItem.getItem_href()
        currentColumn += 1
        
        var = sheet.cell(row = currentRow, column = currentColumn)
        var.value = eachItem.getItemKey()
        currentColumn += 1
        
        var = sheet.cell(row = currentRow, column = currentColumn)
        var.value = eachItem.getItemCategoryName()
        currentColumn += 1
        
        var = sheet.cell(row = currentRow, column = currentColumn)
        var.value = eachItem.getItemCategoryKey()

        currentColumn = 1
        currentRow += 1
        
    
# send subcategory to add to the workbook through.
    if subCatsExist == True:
        index += 1
        for subCategory in currentCategory.getSubCategories():
            addToNewWorkbook(wb, subCategory, currentCategory.getCategoryName(), currentCategory.getCategoryKey(), index = index)

# Look for any empty or blank pages and remove them from the workbook (this usually occurs at the end of the workbook)
    for sheet in wb:
        if sheet.cell(row = 1, column = 1).value == '' or sheet.cell(row = 1, column = 1).value == None:
            wb.remove(sheet)

#*******************************************************************************************************
# Get information from excel for updating purposes.

def loadTree(name):

    # first we pull data for each category and load it into a category object list
    try:
        openName = name + '.xlsx'
        wb = load_workbook(openName)
        nodes = []          # make a list of category objects
        for sheet in wb:
            nodes.append(gatherInformation(sheet))
        wb.close()

        # then we reconstruct our tree using the category's parent key.
        # first we get the root node
        currentParentKey = -1
        root = GroceryCategoryTreeClass.GroceryCategoryTree()
        currentIndex = -1
        for node in nodes:
            currentIndex += 1
            if node.getCategoryParentKey() == currentParentKey:
                root = node
                root.setCategoryParent(None)
                break
        nodes.pop(currentIndex)   # pop the root node from the tree to reduce iteration repetition

        currentCategory = root

        while len(nodes) > 0:
            currentIndex = -1
            childFound = False
            for node in nodes:
                currentIndex += 1
                if node.getCategoryParentKey() == currentCategory.getCategoryKey():
                    node.setCategoryParent(currentCategory)
                    currentCategory.addSubCategory(node)
                    currentCategory = node
                    childFound = True
                    break
            if childFound == True:
                nodes.pop(currentIndex)
            else:
                currentCategory = currentCategory.getCategoryParent()
        return root
        print('\nLoading complete!')
            
    except:
        print('\nFile not found: ' + openName)

def gatherInformation(st):
    tempCategory = GroceryCategoryTreeClass.GroceryCategoryTree()
    tempCategory.setCategoryName(st.cell(row = 1, column = 2).value)
    tempCategory.setCategoryKey(st.cell(row = 2, column = 2).value)
    # note we skip setting category parent, we use parent's Key to find parent object in array of objects
    tempCategory.setCategoryParentKey(st.cell(row = 4, column = 2).value)
    tempCategory.setCategory_href(st.cell(row = 5, column = 2).value)
    tempCategory.setSubCategoriesExist(st.cell(row = 6, column = 2).value)

    currentColumn = 1
    currentRow = 11
    currentCell = st.cell(row = currentRow, column = currentColumn)
    while currentCell.value != '' and currentCell.value != None:
        tempItem = GroceryItemsInformationClass.GroceryItemsInformation()

        value = currentCell.value
        tempItem.setItemName(value)
        currentColumn += 1
        currentCell = st.cell(row = currentRow, column = currentColumn)
        
        value = currentCell.value
        if value == '' or value == None:
            value = 0.0
        tempItem.setItemVariantPrice(float(value))
        currentColumn += 1
        currentCell = st.cell(row = currentRow, column = currentColumn)
        
        value = currentCell.value
        if value == '' or value == None:
            value = 'NA'
        tempItem.setItemVariant(value)
        currentColumn += 1
        currentCell = st.cell(row = currentRow, column = currentColumn)
        
        value = currentCell.value
        if value == '' or value == None:
            value = 0.0
        tempItem.setItemVariantPackSize(float(value))
        currentColumn += 1
        currentCell = st.cell(row = currentRow, column = currentColumn)
        
        value = currentCell.value
        if value == '' or value == None:
            value = 0.0
        tempItem.setItemVariantAltPrice(float(value))
        currentColumn += 1
        currentCell = st.cell(row = currentRow, column = currentColumn)
        
        value = currentCell.value
        if value == '' or value == None:
            value = 'NA'
        tempItem.setItemVariantAlt(value)
        currentColumn += 1
        currentCell = st.cell(row = currentRow, column = currentColumn)
        
        value = currentCell.value
        if value == '' or value == None:
            value = 0.0
        tempItem.setItemUOMPrice(float(value))
        currentColumn += 1
        currentCell = st.cell(row = currentRow, column = currentColumn)
        
        value = currentCell.value
        if value == '' or value == None:
            value = 'NA'
        tempItem.setItemUOM(value)
        currentColumn += 1
        currentCell = st.cell(row = currentRow, column = currentColumn)
        
        value = currentCell.value
        if value == '' or value == None:
            value = 'NA'
        tempItem.setItemBrand(value)
        currentColumn += 1
        currentCell = st.cell(row = currentRow, column = currentColumn)
        
        value = currentCell.value
        tempItem.setItem_href(value)
        currentColumn += 1
        currentCell = st.cell(row = currentRow, column = currentColumn)
        
        value = currentCell.value
        tempItem.setItemKey(value)
        currentColumn += 1
        currentCell = st.cell(row = currentRow, column = currentColumn)
        
        value = currentCell.value
        tempItem.setItemCategoryName(value)
        currentColumn += 1
        currentCell = st.cell(row = currentRow, column = currentColumn)
        
        value = currentCell.value
        if value == '' or value == None:
            value = -1
        tempItem.setItemCategoryKey(int(value))

        tempCategory.addCategoryItem(tempItem)
        currentRow += 1
        currentColumn = 1
        currentCell = st.cell(row = currentRow, column = currentColumn)
        
    # we will get sub categories of current categories after all category objects have been built.
    # we will use parent names to rebuild our tree.
       

    return tempCategory


#*************************************************************************************************************
## these sets of functions web scrapes heb.com in each category and loads items to their respective categories
## in the category tree.

def startGetCategoryItems(root, site):

    sinceEpoch = time.time()
    startTimeObj = time.localtime(sinceEpoch)
    driver = webdriver.Chrome(ChromeDriverManager().install()) 
    driver.get(site)
    time.sleep(3)

    runGetCategoryItems(root, driver, site)
    
    driver.close()
    driver.quit()
    
    sinceEpoch = time.time()
    endTimeObj = time.localtime(sinceEpoch)
    
    print('Started: %d:%d' %(startTimeObj.tm_hour, startTimeObj.tm_min))
    print('Finished: %d:%d' %(endTimeObj.tm_hour, endTimeObj.tm_min))

    return root

def runGetCategoryItems(currentCategory, driver, site):

    
    if currentCategory.doSubCategoriesExist() == False:
        currentCategory = getCategoryItems(currentCategory, driver, site)  # we will pass through the category object and add to it in the callee function
#        printCategoryItemNames(currentCategory)
    else:
        for subCategory in currentCategory.getSubCategories():
            runGetCategoryItems(subCategory, driver, site)
    

def getCategoryItems(currentCategory, dr, site):
    doneScraping = False
    currentSite = site + currentCategory.getCategory_href()
    dr.execute_script("window.open(''); ")
    dr.switch_to.window(dr.window_handles[1])    
    endOfPages = False    
    # while a next button is available

    print('\nScraping ' + currentCategory.getCategoryName())
    itemCount = 0
    while endOfPages == False:
        dr.get(currentSite)
        time.sleep(5)
        html = BeautifulSoup(dr.page_source, 'html.parser')
        listOfItemsOnPage = html.findAll('li', {'class': 'responsivegriditem product-grid-large-fifth product-grid-small-6'})
        print('number of items on current page: ' + str(len(listOfItemsOnPage)))
        for li in listOfItemsOnPage:
            tempItem = GroceryItemsInformationClass.GroceryItemsInformation()
            # this is were we collect the data from each <li> tag for each item on the page and push them to the category's list of items
            # some data are in script tags, while others are in varying types of tags.
            # first we extract bulk of the data from script tags
            scriptTags = li.findAll('script')
            targetScript = scriptTags[0]
                
            try:
                productInfo = json.loads("".join(targetScript.contents), strict=False)
                value = productInfo['name']
                tempItem.setItemName(value)
                value = productInfo['id']
                tempItem.setItemKey(value)
                value = productInfo['brand']
                tempItem.setItemBrand(value)
                value = productInfo['price']
                tempItem.setItemVariantPrice(float(value))
                value = productInfo['variant']
                tempItem.setItemVariant(value)
                itemCount += 1
            except:
                key = li.find('form')['data-product-id']
                tempItem.setItemKey(key)
                print(key + ' : unable to get json, item skipped')

            spanTags = li.findAll('span', {'class': 'item-size'})
            if spanTags != []:
                span = spanTags[0]      # usually the first index of the list.
                altV = cleanupUOM(span.get_text())
                altVPackSize = cleanupFloats(span.get_text())
                try:
                    tempItem.setItemVariantPackSize(float(altVPackSize))
                except:
                    tempItem.setItemVariantPackSize(1)
                try:
                    tempItem.setItemVariantAlt(altV)
                except:
                    tempItem.setItemVariantAlt(tempItem.getItemVariant())

                try:
                    altVPrice = tempItem.getItemVariantPrice() / tempItem.getItemVariantPackSize()
                    tempItem.setItemVariantAltPrice(altVPrice)
                except:
                    tempItem.setItemVariantAltPrice(tempItem.getItemVariantPrice())
                

                    
            # then we gather data from other tags            
            # pimary uom and primary uom price will be based off of the <span> tag with class="uomSalePrice"
            # this holds the item's price at uom
            spanTags = li.findAll('span', {'class': 'uomSalePrice'})
            if spanTags != []:
                span = spanTags[0]      # usually the first index of the list.
                uom = cleanupUOM(span.get_text())
                uomPrice = cleanupFloats(span.get_text())
                try:
                    tempItem.setItemUOMPrice(float(uomPrice))
                except:
                    tempItem.setItemUOMPrice(tempItem.getItemVariantPrice())
                try:
                    tempItem.setItemUOM(uom)
                except:
                    tempItem.setItemUOM(tempItem.getItemVariant())


            aTags = li.findAll('a')
            try:
                href = aTags[0]['href']
            except:
                print(tempItem.getItemKey() + ' : ' + tempItem.getItemName() + ' has no aTag[0] with an href.')
            tempItem.setItem_href(href)
            tempItem.setItemCategoryName(currentCategory.getCategoryName())
            tempItem.setItemCategoryKey(currentCategory.getCategoryKey())
            

            # once we have our data collected for the item, we append it to the item list of the category object.
            currentCategory.addCategoryItem(tempItem)

        # after we get the items from teh page loaded, we move on to the next page if a next page exists.
        nextPageFinder = html.findAll('a', {'aria-label': 'go to next page'})
        if nextPageFinder != []:    # if there exists a 'go to next page' <a> tag, we go to the next page
            nxpg = nextPageFinder[0]
            currentSite = site + nxpg['href']
        else:                       # if there does not exist a 'go to next page' <a> tag, we have found all pages in currentCategory
            endOfPages = True


    dr.close()
    dr.switch_to.window(dr.window_handles[0])

    print('\nNumber of items scraped in ' + currentCategory.getCategoryName() + ': ' + str(itemCount))

    return currentCategory


def cleanupUOM(string):

    if string != '' or string != None:
        newString = re.sub('[0-9()\s$/.]', '', string)
        newString = re.sub('avg', '', newString)
        newString = re.sub('bag', '', newString)
    else:
        newString = 'NA'

    return newString

def cleanupFloats(string):

    if string != '' or string != None:
        newString = re.sub('[a-z()\s$/]', '', string)
    else:
        newString = '0.0'

    return newString

def printCategoryItemNames(category):
    print('\n' + category.getCategoryName() + '\n')
    print('Number of Items: ' + str(category.getCategoryItemsCount()))
    for item in category.getCategoryItems():
        print(item.getItemName())
        

#*************************************************************************************************************
## this is for testing, to scrape item information and save it to the workbook.
def loadCheese(name, site):

    print('\nOpening file')
    file = 'D:\\Python Projects\\' + name + '.xlsx'
    wb = load_workbook(file)
    cheese = 419    # the tab that cheese is on in the HEB workbook.
    print('\nGetting cheese tab')
    st = wb.worksheets[cheese]
    temp = GroceryCategoryTreeClass.GroceryCategoryTree()
    temp.setCategoryName(st.cell(row = 1, column = 2).value)
    temp.setCategoryKey(st.cell(row = 2, column = 2).value)
    # note we skip setting category parent, we use parent's Key to find parent object in array of objects
    temp.setCategoryParentKey(st.cell(row = 4, column = 2).value)
    temp.setCategory_href(st.cell(row = 5, column = 2).value)
    temp.setSubCategoriesExist(st.cell(row = 6, column = 2).value)
    wb.close()
    
    print('\nCheese tab loaded')

    # it is safe to assume that non-leaf nodes do not have items to scrape from the website.
    if temp.doSubCategoriesExist() == False:
        driver = webdriver.Chrome(ChromeDriverManager().install()) 
        driver.get(site)
        time.sleep(3)
        
        getCategoryItemsTest(temp, driver, site)  # we will pass through the category object and add to it in the callee function
    
        driver.close()
        driver.quit()

    
    
def getCategoryItemsTest(currentCategory, dr, site):
    doneScraping = False
    currentSite = site + currentCategory.getCategory_href()
    dr.execute_script("window.open(''); ")
    dr.switch_to.window(dr.window_handles[1])    
    endOfPages = False    
    # while a next button is available
    while endOfPages == False:
        dr.get(currentSite)
        time.sleep(5)
        html = BeautifulSoup(dr.page_source, 'html.parser')
        listOfItemsOnPage = html.findAll('li', {'class': 'responsivegriditem product-grid-large-fifth product-grid-small-6'})
        for li in listOfItemsOnPage:
            tempItem = GroceryItemsInformationClass.GroceryItemsInformation()
            # this is were we collect the data from each <li> tag for each item on the page and push them to the category's list of items
            # some data are in script tags, while others are in varying types of tags.
            # first we extract bulk of the data from script tags
            scriptTags = li.findAll('script')
            targetScript = scriptTags[0]
            productInfo = json.loads("".join(targetScript.contents))
            tempItem.setItemName(productInfo['name'])
            tempItem.setItemKey(productInfo['id'])
            tempItem.setItemBrand(productInfo['brand'])
            tempItem.setItemVariantPrice(float(productInfo['price']))
            tempItem.setItemVariant(productInfo['variant'])

            # then we gather data from other tags
            
            # pimary uom and primary uom price will be based off of the <span> tag with class="uomSalePrice"
            # this holds the item's price and uom in terms of weighted ounce as a standard on HEB.com
            spanTags = li.findAll('span', {'class': 'uomSalePrice'})
            if spanTags != []:
                span = spanTags[0]      # usually the first index of the list.
                pUOM = re.sub('[0-9()\s$/.]', '', (span.get_text()))
                pUOMPrice = re.sub('[a-z()\s$/]', '', (span.get_text()))
                tempItem.setItemPrimaryUOMPrice(float(pUOMPrice))
                tempItem.setItemPrimaryUOM(pUOM)
            else:
                print('<span> tag with class uomSalePrice DNE!')


            # variant uom and variant uom size will be based off of the <span> tag with class="item-size"
            # this holds the item's variant UOM and variant UOM size per variant (usually each)
            # example: variant is 1 each which equals variant UOM Size 10 and variant UOM ct
            # 1 each = 10 ct
            spanTags = li.findAll('span', {'class': 'item-size'})
            if spanTags != []:
                span = spanTags[0]      # usually the first index of the list.
                vUOM = re.sub('[0-9()\s$/.]', '', (span.get_text()))
                vUOMSize = re.sub('[a-z()\s$/]', '', (span.get_text()))
                tempItem.setItemVariantUOMSize(float(vUOMSize))
                tempItem.setItemVariantUOM(vUOM)

                # secondary uom and uom price will give a cost per variant uom
                # variant price / variant UOM size
                sUOM = vUOM
                sUOMPrice = float(productInfo['price']) / float(vUOMSize)
                tempItem.setItemSecondaryUOMPrice(float(sUOMPrice))
                tempItem.setItemSecondaryUOM(sUOM)
                
            else:
                print('<span> tag with class uomSalePrice DNE!')

            aTags = li.findAll('a')
            href = aTags[0]['href']
            tempItem.setItem_href(href)
            tempItem.setItemCategoryName(currentCategory.getCategoryName())
            tempItem.setItemCategoryKey(currentCategory.getCategoryKey())

            # once we have our data collected for the item, we append it to the item list of the category object.
            currentCategory.addCategoryItem(tempItem)

        # after we get the items from teh page loaded, we move on to the next page if a next page exists.
        nextPageFinder = html.findAll('a', {'aria-label': 'go to next page'})
        if nextPageFinder != []:    # if there exists a 'go to next page' <a> tag, we go to the next page
            nxpg = nextPageFinder[0]
            currentSite = site + nxpg['href']
        else:                       # if there does not exist a 'go to next page' <a> tag, we have found all pages in currentCategory
            endOfPages = True


    dr.close()
    dr.switch_to.window(dr.window_handles[0])


#**********************************************************************************************************************************
# function to save all items from HEB item tree to a single page csv file.

def saveItemsInTree(root, name):
    wb = Workbook()

    wb.active = 0
    sheet = wb.active
    label = sheet.cell(row = 1, column = 1)
    label.value = 'Name'       
    label = sheet.cell(row = 1, column = 2)
    label.value = 'Variant Price'       
    label = sheet.cell(row = 1, column = 3)
    label.value = 'Variant'       
    label = sheet.cell(row = 1, column = 4)
    label.value = 'Variant Pack Size'       
    label = sheet.cell(row = 1, column = 5)
    label.value = 'Variant Alt Price'       
    label = sheet.cell(row = 1, column = 6)
    label.value = 'Variant Alt'       
    label = sheet.cell(row = 1, column = 7)
    label.value = 'UOM Price'       
    label = sheet.cell(row = 1, column = 8)
    label.value = 'UOM'       
    label = sheet.cell(row = 1, column = 9)
    label.value = 'Brand'   
    label = sheet.cell(row = 1, column = 10)
    label.value = 'href'       
    label = sheet.cell(row = 1, column = 11)
    label.value = 'Item key'       
    label = sheet.cell(row = 1, column = 12)
    label.value = 'Category Name'       
    label = sheet.cell(row = 1, column = 13)
    label.value = 'Category Key'
    label = sheet.cell(row = 1, column = 12)
    label.value = 'Parent Category Name'       
    label = sheet.cell(row = 1, column = 13)
    label.value = 'Parent Category Key'

    currentRow = 2
    currentRow = saveCurrentCategoryItemsToWorkbook(wb, sheet, currentRow, root)
            
    openFile = name + '.xlsx'
    wb.save(openFile)

def saveCurrentCategoryItemsToWorkbook(wb, sheet, currentRow, currentCategory):
    # if the current category does not have a sub category, then  we know we have items to save to the workbook.
    # this is where we loop each product item from currentCategory
    # and their respective product information.
    currentColumn = 1
    if len(currentCategory.getCategoryItems()) > 0:
        for eachItem in currentCategory.getCategoryItems():
            var = sheet.cell(row = currentRow, column = currentColumn)
            var.value = eachItem.getItemName()
            currentColumn += 1
                
            var = sheet.cell(row = currentRow, column = currentColumn)
            var.value = eachItem.getItemVariantPrice()
            currentColumn += 1
                
            var = sheet.cell(row = currentRow, column = currentColumn)
            var.value = eachItem.getItemVariant()
            currentColumn += 1
            
            var = sheet.cell(row = currentRow, column = currentColumn)
            var.value = eachItem.getItemVariantPackSize()
            currentColumn += 1
                
            var = sheet.cell(row = currentRow, column = currentColumn)
            var.value = eachItem.getItemVariantAltPrice()
            currentColumn += 1
                
            var = sheet.cell(row = currentRow, column = currentColumn)
            var.value = eachItem.getItemVariantAlt()
            currentColumn += 1
                
            var = sheet.cell(row = currentRow, column = currentColumn)
            var.value = eachItem.getItemUOMPrice()
            currentColumn += 1
                
            var = sheet.cell(row = currentRow, column = currentColumn)
            var.value = eachItem.getItemUOM()
            currentColumn += 1
                
            var = sheet.cell(row = currentRow, column = currentColumn)
            var.value = eachItem.getItemBrand()
            currentColumn += 1
                
            var = sheet.cell(row = currentRow, column = currentColumn)
            var.value = eachItem.getItem_href()
            currentColumn += 1
                
            var = sheet.cell(row = currentRow, column = currentColumn)
            var.value = eachItem.getItemKey()
            currentColumn += 1
                
            var = sheet.cell(row = currentRow, column = currentColumn)
            var.value = eachItem.getItemCategoryName()
            currentColumn += 1
                
            var = sheet.cell(row = currentRow, column = currentColumn)
            var.value = eachItem.getItemCategoryKey()
            currentColumn += 1
                
            var = sheet.cell(row = currentRow, column = currentColumn)
            var.value = currentCategory.getCategoryParent().getCategoryName()
            currentColumn += 1
                
            var = sheet.cell(row = currentRow, column = currentColumn)
            var.value = currentCategory.getCategoryParentKey()

            currentColumn = 1
            currentRow += 1
           
    # recursively go through all sub categories until we reach a category that does not have sub categories.
    if currentCategory.doSubCategoriesExist() == True:
        for subCategory in currentCategory.getSubCategories():
            currentRow = saveCurrentCategoryItemsToWorkbook(wb, sheet, currentRow, subCategory)

    return currentRow
    
    

#***********************************************************************************************************************************
# menu options.


def treeMenuOptions():
    
    choice = -1
    while choice < 0 or choice > 6:
        print('Beer advocate scraper, will be used for collecting our data')
        print('What operation are we running?')
        print('0. QUIT')
        print('1. Build Tree: scrape grocery website for its CATEGORIES to build a tree')
        # NOTE: only if a categories tree already exists in the program!
        print('2. Print Tree: print the tree to screen')
        print('3. Save Tree: save the tree to an excel file')
        # NOTE: requires loading cateogires and items tree from excel
        print('4. Load Tree: load the tree from an excel file')
        print('5. Get Category Items: scrape grocery website for items in each category\n   and add to the tree')
        print('6. Save all times to one csv file')    
#        print('6. Look for missing CATEGORIES in Excel file')
#        print('7. Load and print CATEGORIES tree')

# need a try except here.
        try:
            choice = int(input())
        except:
            print('\nInvalid entry: value entered for menu choice not a valid option!\n')
            print(choice)

        
    return choice

# option 1
# function that builds a category tree
def buildCategoryTree(gURL, gName):
    print('\nBuilding category tree..\n')
    # our tree object where we append the nodes and their respective data values.
    tree = GroceryCategoryTreeClass.GroceryCategoryTree()
    # gets the html via selenium and beautifulsoup combination
    soupSource = seleniumGetsHTML(gURL)
    # get the tree's starting node.
    tree = buildTree(soupSource, gURL)
    print('\nTree has been built!')
    return tree

# option 2
# function that prints a visual representation of the tree to the screen
def printTree(currentCategory):
    print('\nPrinting category tree to screen..\n')
    print(currentCategory.getCategoryName())
    printCategoryTree(currentCategory, 1)
    print('\nTree has been printed!')

# option 3
# function to save tree to a workbook
def saveToWorkbook(tree, name):

    print('\nSaving to workbook titled ' + name + '..')
    createWorkbook(tree, name)
    print('\nSaving complete!')


# option 4
# function to save tree to a workbook
def loadFromWorkbook(name):

    print('\nLoading tree from workbook titled ' + name + '..')
    tree = loadTree(name)
    return tree

# option 5
# function to add items to each category in the tree
# usually the leaf nodes of a tree
def scrapeCategoryItems(tree, gURL):

    print('\nGathering item information from website ' + gURL + '..')
    tree = startGetCategoryItems(tree, gURL)
    print('\nGathering of item information Complete!')
    return tree


# OPTION FOR TESTING PURPOSES
def testOption(tree):
    print('\nBeginning Test')
    saveItemsInTree(tree, 'HEB_Items')
    print('\nTest Completed')
                        
#*******************************************************************************************************
# main function    
def main():

    # our tree object where we append the nodes and their respective data values.
    groceryCategoryTreeObject = GroceryCategoryTreeClass.GroceryCategoryTree()

#    groceryChoice
#    while groceryChoice != 1:
#        print('Grocery Web Scraping App')
#        print('Select the grocery you want to run the app on')
#        print('1. HEB')
#        choice = input()

    # THIS ONLY WORKS FOR HEB.COM At the moment
    # page we will be visiting to build our tree
    groceryURL = 'https://www.heb.com'
    # Grocery Name
    groceryName = 'HEB'
    operationChoice = treeMenuOptions()

    while(operationChoice != 0):
        if(operationChoice == 1):
            groceryCategoryTreeObject = buildCategoryTree(groceryURL, groceryName)
            operationChoice = treeMenuOptions()            
            
        elif(operationChoice == 2):
            printTree(groceryCategoryTreeObject)
            operationChoice = treeMenuOptions()
            
        elif(operationChoice == 3):
            saveToWorkbook(groceryCategoryTreeObject, groceryName)
            operationChoice = treeMenuOptions()            
            
        elif(operationChoice == 4):
            groceryCategoryTreeObject = loadFromWorkbook(groceryName)
            operationChoice = treeMenuOptions()  
            
        elif(operationChoice == 5):
            scrapeCategoryItems(groceryCategoryTreeObject, groceryURL)
            operationChoice = treeMenuOptions()
            
        elif(operationChoice == 6):
            testOption(groceryCategoryTreeObject)
            operationChoice = treeMenuOptions()
            
    #    elif(operationChoice == 6):
            
    #    elif(operationChoice == 7):

        
        
main()
